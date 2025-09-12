from __future__ import annotations
import argparse, os, re
from openpyxl import load_workbook

def sanitize(name: str) -> str:
    name = re.sub(r'[<>:"/\\|?*\x00-\x1F]+', " ", name)
    name = re.sub(r"\s+", " ", name).strip().rstrip(". ")
    return name[:120]

def xlsx_to_txt(path: str) -> str:
    wb = load_workbook(path, data_only=True, read_only=True)
    base = sanitize(os.path.splitext(os.path.basename(path))[0])
    os.makedirs("data/raw", exist_ok=True)
    out_path = os.path.join("data", "raw", f"{base}.txt")

    with open(out_path, "w", encoding="utf-8") as out:
        for ws in wb.worksheets:
            out.write(f"# Sheet: {ws.title}\n")
            for row in ws.iter_rows(values_only=True):
                cells = [(str(c) if c is not None else "") for c in row]
                out.write("\t".join(cells) + "\n")
            out.write("\n")
    return out_path

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--path", required=True, help="Caminho do arquivo .xlsx")
    args = ap.parse_args()
    out = xlsx_to_txt(args.path)
    print(f"Salvo: {out}")
